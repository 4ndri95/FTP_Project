import openpyxl
import os
import pandas as pd
from datetime import datetime
import re
import cx_Oracle
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Inicializa o cliente Oracle
cx_Oracle.init_oracle_client(lib_dir=r"//path//to//sql")

# Configuração do logging
logging.basicConfig(filename='pdf_collection.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def collect_pdfs(directories):
    pdf_files = []
    for directory in directories:
        try:
            logging.info(f"Processando diretório: {directory}")
            if not os.path.isdir(directory):
                logging.error(f"Diretório não encontrado: {directory}")
                continue

            files_in_directory = os.listdir(directory)
            if not files_in_directory:
                logging.warning(f"O diretório {directory} está vazio.")
                continue 

            for file in files_in_directory:
                if file.endswith('.pdf'):
                    cleaned_uc = re.sub(r'\D', '', file)
                    pdf_files.append(cleaned_uc)
                    logging.info(f"Arquivo PDF encontrado: {file}")
                else:
                    logging.warning(f"Arquivo ignorado (não é PDF): {file}")

            logging.info(f"Processamento do diretório {directory} concluído com sucesso.")
        except FileNotFoundError:
            logging.error(f"Diretório não encontrado: {directory}")
        except Exception as e:
            logging.error(f"Erro ao processar o diretório {directory}: {e}")
    return pdf_files

def get_current_date(date_format="%d.%m.%Y"):
    return datetime.now().strftime(date_format)

def connect_to_database(config, retries=3):
    dsn = cx_Oracle.makedsn(config['address'], config['port'], service_name=config['service_name'])
    for attempt in range(retries):
        try:
            return cx_Oracle.connect(config['user'], config['password'], dsn)
        except cx_Oracle.DatabaseError as e:
            logging.error(f"Erro ao conectar ao banco de dados: {e}. Tentativa {attempt + 1} de {retries}.")
            if attempt < retries - 1:
                continue  
            else:
                raise  

def execute_query(cursor, ucs):
    sql = f"""
        SELECT DISTINCT
        a.cod_un_cons_uee,
        a.cod_loc_uee
        FROM 
            rededes.cad_uc_ee a
        WHERE
             a.cod_un_cons_uee IN {ucs}
    """
    try:
        cursor.execute(sql)
        return cursor.fetchall()
    except cx_Oracle.DatabaseError as e:
        logging.error(f"Erro ao executar a consulta SQL: {e}")
        return []

def save_to_excel(results, current_date):
    if results:
        results_df = pd.DataFrame(results, columns=['cod_un_cons_uee', 'cod_loc_uee'])
        results_df['Data'] = current_date

        spreadsheet_name = f"Processes retrieved from FTP on {current_date}.xlsx"
        results_df.to_excel(spreadsheet_name, index=False, sheet_name='SQL_Results')

        wb = Workbook(spreadsheet_name)
        ws = wb.active

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        for row in range(1, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row == 1:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = openpyxl.styles.PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

        wb.save(spreadsheet_name)
        print(f"Planilha criada com formatação: {spreadsheet_name}")
    else:
        print("Nenhum resultado encontrado para a consulta SQL.")

def main():
    directories = [
        '//path//to//directory',
        '//path//to//directory',
        '//path//to//directory',
        '//path//to//directory'
    ]

    pdf_files = collect_pdfs(directories)

    ucs = "','".join(pdf_files)
    ucs = f"('{ucs}')"

    date_format = "%d.%m.%Y"
    current_date = get_current_date(date_format)

    db_config = {
        'user': 'User',
        'password': 'Password',
        'address': 'address',
        'port': port,
        'service_name': 'service_name'
    }

    try:
        with connect_to_database(db_config) as conn:
            with conn.cursor() as cursor:
                results = execute_query(cursor, ucs)

        save_to_excel(results, current_date)

    except cx_Oracle.DatabaseError as e:
        logging.error(f"Erro ao conectar ou executar a consulta no banco de dados: {e}")
        print("Ocorreu um erro ao tentar acessar o banco de dados. Verifique os logs para mais detalhes.")
    except Exception as e:
        logging.error(f"Erro inesperado: {e}")
        print("Ocorreu um erro inesperado. Verifique os logs para mais detalhes.")

if __name__ == "__main__":
    main()
    
